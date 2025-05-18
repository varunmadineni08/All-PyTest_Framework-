import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By


def get_data():
    final_list=[]
    workbook=openpyxl.load_workbook("PytestPackage/Book.xlsx")
    sheet=workbook["Sheet1"]
    total_rows=sheet.max_row
    total_cols=sheet.max_column
    for r in range(2,total_rows+1):
        row_list=[]
        for c in range(1,total_cols+1):
            cell_value=sheet.cell(row=r,column=c).value
            row_list.append(cell_value if cell_value is not None else "")
        final_list.append(row_list)
    return final_list






@pytest.mark.parametrize("username,password",get_data())
def test_login(username,password):
    driver=webdriver.Edge()
    driver.get("https://www.saucedemo.com/")
    driver.maximize_window()
    driver.find_element(By.ID,'user-name').send_keys(username)
    driver.find_element(By.ID, 'password').send_keys(password)
    driver.find_element(By.ID,'login-button').click()
    driver.quit()
